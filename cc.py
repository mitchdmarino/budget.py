import os
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

directory_path = './cc_statements_USBANK'
cc_statements = [f for f in os.listdir(directory_path) if f.endswith('.pdf')]
charges_list = []
credits_list = []

# Define a function to extract the year from the PDF text
def extract_year(text):
    year_pattern = r'(\d{4})'  # Adjust this pattern based on how the year is presented in the PDF
    match = re.search(year_pattern, text)
    if match:
        print("''''''''''''''''''''''''''''DATE YEAR ", match.group(1))
        return match.group(1)
    return None

# Function to add the year to a date string
def add_year(date_str, year):
    date_format = '%m/%d'
    full_date_str = f"{date_str}/{year}"
    return datetime.strptime(full_date_str, f"{date_format}/%Y")

# Define a function to extract the section text
def extract_section(text, start_title, end_title=None):
    pattern = re.compile(fr'{re.escape(start_title)}(.*?){re.escape(end_title) if end_title else ""}', re.DOTALL)
    match = pattern.search(text)
    if match:
        print(start_title)
        print(match.group(1))
        print(match.group(0))
        return match.group(1)
    return ""

# Define a function to parse the text for itemized entries
def parse_entries(text):
    # Adjust the pattern according to the format of your entries
    # Example pattern for: mm/dd mm/dd 1234 Description $123.45
    pattern = r'(\d{2}/\d{2})\s+(\d{2}/\d{2})\s+(\d{4})\s+(.+?)\s+(-?\$[\d,]+\.\d{2})'
    matches = re.findall(pattern, text)
    return matches

# Open the PDF file
def parse_files(path):
    path = directory_path + "/" + path
    with pdfplumber.open(path) as pdf:
        all_pages_text = []
        
        # Loop through all pages and extract text
        for page in pdf.pages:
            text = page.extract_text()
            all_pages_text.append(text)

    # Combine all the text into a single string
    combined_text = "\n".join(all_pages_text)

    # Extract the year from the combined text
    year = extract_year(combined_text)
    if not year:
        raise ValueError("Year not found in the PDF.")

    # Extract sections for charges and credits
    charges_text = extract_section(combined_text, "Purchases and Other Debits", "TOTAL THIS PERIOD")
    credits_text = extract_section(combined_text, "Payments and Other Credits", "TOTAL THIS PERIOD")

    # Parse the sections for itemized entries
    charges = parse_entries(charges_text)
    credits = parse_entries(credits_text)

    # Convert the list of charges into a pandas DataFrame
    df_charges = pd.DataFrame(charges, columns=['Post Date', 'Transaction Date', 'Ref#', 'Description', 'Amount'])
    df_credits = pd.DataFrame(credits, columns=['Post Date', 'Transaction Date', 'Ref#', 'Description', 'Amount'])

    # Clean and format the dataframes as needed
    for df in [df_charges, df_credits]:
        df['Post Date'] = df['Post Date'].apply(lambda x: add_year(x, year))
        df['Transaction Date'] = df['Transaction Date'].apply(lambda x: add_year(x, year))
        df['Amount'] = df['Amount'].replace({'\$': '', ',': ''}, regex=True).astype(float)

    # Print the dataframes
    print("Charges:")
    print(df_charges)
    charges_list.append(df_charges)
    
    print("\nCredits:")
    print(df_credits)
    credits_list.append(df_credits)


for statement in cc_statements: 
    parse_files(statement)

# Concatenate all charges and credits DataFrames
print("################################################################################################")
print(charges_list)
print(credits_list)
all_charges_df = pd.concat(charges_list, ignore_index=True)
all_credits_df = pd.concat(credits_list, ignore_index=True)

# Save to CSV
all_charges_df.to_csv('./organized_cc_statements/charges.csv', index=False)
all_credits_df.to_csv('./organized_cc_statements/credits.csv', index=False)

print("Charges and credits have been successfully combined and saved to CSV files.")


####################################################################################
# Now we need to add the data to excel

def import_csv_to_dataframe(csv_file): 
    try: 
        df = pd.read_csv(csv_file)
        return df
    except FileNotFoundError:
        print(f"CSV file '{csv_file}' not found.")
        return None
    except Exception as e:
        print(f"Error occurred while importing '{csv_file}': {e}")
        return None
    
# Function to save DataFrame to Excel
def save_to_excel(excel_file, sheet_name, df):
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Use mode='a' to append to existing file if it exists
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            def formatDates():
                wb = load_workbook(excel_file)

                # Select the specific sheet
                
                ws = wb[sheet_name]

                # Specify the range of cells (column A in this example) to convert to dates
                start_row = 2  # Assuming data starts from row 2 (excluding header)
                end_row = ws.max_row  # Assuming data extends to the last row
                date_column = 'A'  # Column A

                # Convert text values to dates in the specified column
                date_format = NamedStyle(name='date_format', number_format='MM/DD/YYYY')
                for row in range(start_row, end_row + 1):
                    cell = ws[date_column + str(row)]
                    if isinstance(cell.value, str):
                        try:
                            date_value = datetime.strptime(cell.value, '%m/%d/%Y')  # Adjust format as needed
                            print(date_value, "\n")
                            ws[cell.coordinate] = date_value
                            ws[cell.coordinate].style = date_format 
                            
                        except ValueError:
                            pass  # Handle invalid date formats or skip errors

                # Save the modified workbook
                wb.save(excel_file)
            #formatDates()
        print(f"Data from '{sheet_name}' imported into '{excel_file}'")
    except Exception as e:
        print(f"An error occurred while saving to '{excel_file}': {e}")

# Directory containing CSV files
csv_dir = './organized_cc_statements'
excel_file = 'Budget.xlsx'  # Name of the Excel file to create or append to

# List CSV files in the directory
csv_files = [os.path.join(csv_dir, f) for f in os.listdir(csv_dir) if f.endswith('.csv')]

# Process each CSV file
for csv_file in csv_files:
    df = import_csv_to_dataframe(csv_file)
    if df is not None:
        sheet_name = "raw_cc_charges"  
        save_to_excel(excel_file, sheet_name, df)


# now add the dates formula 
def format_dates(): 
    # Load the existing Excel workbook
    wb = load_workbook(excel_file)

    # Select the specific sheet
    sheet_name = 'raw_cc_charges'  # Replace with your sheet name
    ws = wb[sheet_name]

    # Insert a new column with a formula (e.g., sum of columns A and B)
    new_column_index = 6  # Insert new column after column B (index 2)
    formula = '=DATEVALUE(A{row_num})'

    # Insert header for the new column
    header_cell = ws.cell(row=1, column=new_column_index, value='Date as date')


    # Apply formula to each cell in the new column
    for row in range(2, ws.max_row + 1):
        formula_cell = ws.cell(row=row, column=new_column_index, value=formula.format(row_num=row))

    # Adjust column widths if needed
    ws.column_dimensions[get_column_letter(new_column_index)].width = 15

    # Create a table (optional)
    table = Table(displayName="Table2", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the modified workbook
    wb.save(excel_file)

format_dates()