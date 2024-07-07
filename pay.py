import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def extract_tables_from_pdf(pdf_file, page_number=0):
    with pdfplumber.open(pdf_file) as pdf:
        page = pdf.pages[page_number]
        
        # Extract all tables from the page
        tables = page.extract_tables()

    return tables

def extract_relevant_info(table, extraction_data):    
    # Convert table to DataFrame
    if 'Check Date' in table[0]:
        for row in table[1:]: 
            date = row[-2].strip()
            #date = pd.to_datetime(date)
            extraction_data['Pay Date'] = date
    if 'Gross Pay' in table[0]: 
        row = table[1]
        
        hours = row[1].strip()
        hours = hours.replace(',', '')
        extraction_data['Hours'] = float(hours)

        gross = row[2].strip()
        gross = gross.replace(',', '')
        extraction_data['Gross Pay'] = float(gross)

        pre = row[3].strip()
        pre.replace(',', '')
        extraction_data['Pre Tax Deductions'] =float(pre)

        statutory = row[4].strip()
        statutory = statutory.replace(',', '')
        extraction_data['Statutory Taxes'] = float(statutory)

        post = row[5].strip()
        post = post.replace(',', '')
        post = float(post)
        if post < 0:
            post = 0
        extraction_data['Post Tax Deductions'] = float(post)

        net = row[6].strip()
        net = net.replace(',', '')
        extraction_data['Net Pay'] = float(net)
        
        return extraction_data
    
pay_files = [os.path.join("./paystubs", f) for f in os.listdir('./paystubs') if f.endswith('.pdf')]
print(len(pay_files))
dfs = []
# Extract tables from a sample pay stub
for stub in pay_files: 
    tables = extract_tables_from_pdf(stub)
    extraction_data = {
        'Pay Date': None,
        'Gross Pay': None, 
        'Net Pay': None,
        'Hours': None,
        'Pre Tax Deductions': None,
        'Post Tax Deductions': None, 
    }
    # Convert the tables to DataFrame (adjust if you have multiple tables)
    for table in tables: 
        extract_relevant_info(table, extraction_data)
    df = pd.DataFrame([extraction_data])
    print(df)
    dfs.append(df)

# Convert 'Pay Date' column to datetime format in each DataFrame
#for df in dfs:
    #df['Pay Date'] = pd.to_datetime(df['Pay Date'], format='%m/%d/%Y')

# Convert numeric columns to numeric format, coercing errors to NaN in each DataFrame
'''
numeric_columns = ['Gross Pay', 'Net Pay', 'Hours', 'Post Tax Deductions', 'Pre Tax Deductions', 'Statutory Taxes']
for df in dfs:
    for col in numeric_columns:
        try:
            df['Numeric_Text'] = pd.to_numeric(df['Numeric_Text'])
        except ValueError as e:
            print(f"Error converting column: {e}")
        df[col] = pd.to_numeric(df[col], errors='coerce')
'''
# Concatenate DataFrames in dfs list into one DataFrame
merged_df = pd.concat(dfs, ignore_index=True)

# Export merged DataFrame to CSV file
merged_df.to_csv('./organized_pay_stubs/merged_paystubs.csv', index=False)




####################################################################################
# Now we need to add the data to excel

def import_csv_to_dataframe(csv_file): 
    try: 
        df = pd.read_csv(csv_file)
        return df
    except FileNotFoundError:
        print(f"CSV file '{csv_file}' not found.")
        return None
    except Exception as e:
        print(f"Error occurred while importing '{csv_file}': {e}")
        return None
    
# Function to save DataFrame to Excel
def save_to_excel(excel_file, sheet_name, df):
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Use mode='a' to append to existing file if it exists
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            def formatDates():
                wb = load_workbook(excel_file)

                # Select the specific sheet
                
                ws = wb[sheet_name]

                # Specify the range of cells (column A in this example) to convert to dates
                start_row = 2  # Assuming data starts from row 2 (excluding header)
                end_row = ws.max_row  # Assuming data extends to the last row
                date_column = 'A'  # Column A

                # Convert text values to dates in the specified column
                date_format = NamedStyle(name='date_format', number_format='MM/DD/YYYY')
                for row in range(start_row, end_row + 1):
                    cell = ws[date_column + str(row)]
                    if isinstance(cell.value, str):
                        try:
                            date_value = datetime.strptime(cell.value, '%m/%d/%Y')  # Adjust format as needed
                            print(date_value, "\n")
                            ws[cell.coordinate] = date_value
                            ws[cell.coordinate].style = date_format 
                            
                        except ValueError:
                            pass  # Handle invalid date formats or skip errors

                # Save the modified workbook
                wb.save(excel_file)
            #formatDates()
        print(f"Data from '{sheet_name}' imported into '{excel_file}'")
    except Exception as e:
        print(f"An error occurred while saving to '{excel_file}': {e}")

# Directory containing CSV files
csv_dir = './organized_pay_stubs'
excel_file = 'Budget.xlsx'  # Name of the Excel file to create or append to

# List CSV files in the directory
csv_files = [os.path.join(csv_dir, f) for f in os.listdir(csv_dir) if f.endswith('.csv')]

# Process each CSV file
for csv_file in csv_files:
    df = import_csv_to_dataframe(csv_file)
    if df is not None:
        sheet_name = "raw_income"  # Use filename as sheet name
        save_to_excel(excel_file, sheet_name, df)


# now add the dates formula 
def format_dates(): 
    # Load the existing Excel workbook
    wb = load_workbook(excel_file)

    # Select the specific sheet
    sheet_name = 'raw_income'  # Replace with your sheet name
    ws = wb[sheet_name]

    # Insert a new column with a formula (e.g., sum of columns A and B)
    new_column_index = 8  # Insert new column after column B (index 2)
    formula = '=DATEVALUE(A{row_num})'

    # Insert header for the new column
    header_cell = ws.cell(row=1, column=new_column_index, value='Date as date')


    # Apply formula to each cell in the new column
    for row in range(2, ws.max_row + 1):
        formula_cell = ws.cell(row=row, column=new_column_index, value=formula.format(row_num=row))

    # Adjust column widths if needed
    ws.column_dimensions[get_column_letter(new_column_index)].width = 15

    # Create a table (optional)
    table = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the modified workbook
    wb.save(excel_file)

format_dates()